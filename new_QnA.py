


import pyttsx3
import speech_recognition as sr
import openpyxl
import streamlit as st

# Set Streamlit page title and favicon
st.set_page_config(page_title="Exam", page_icon="📝")

# Initialize the engine
engine = pyttsx3.init()

# Set properties (optional)
engine.setProperty('rate', 150)    # Speed percent (can go over 100)
engine.setProperty('volume', 0.8)  # Volume 0-1

# Initialize the recognizer
r = sr.Recognizer()

# List of questions
questions = [    ["What is the capital of India?",["Option 1: New Delhi", "Option 2: Mumbai", "Option 3: Bangalore", "Option 4: Chennai"]],
    ["What is the largest country in the world by area?",["Option 1: Russia", "Option 2: Canada", "Option 3: China", "Option 4: USA"]],
    ["What is the currency of Japan?",["Option 1: Yen", "Option 2: Rupee", "Option 3: Dollar", "Option 4: Euro"]]
]

# Initialize the Excel sheet
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Q&A"
sheet['A1'] = "Question"
sheet['B1'] = "Answer"
row = 2
lo=1
# Set the Streamlit app header and description
st.title("Exam")
st.markdown("Click the button below to start the exam.")

# Start the exam when the user clicks the "Start Exam" button
if st.button("Start Exam"):
    # Take up to 10 questions from the list, convert them to speech, listen to the answers, and save them in the Excel sheet
    for i in range(min(3, len(questions))):
        # Get the question from the list and convert it to speech
        question = questions[i][0]
        options = questions[i][1]
        # Display the question on the Streamlit page
        st.markdown(f"### Q{i+1}: {question}")
        for option in options:
            st.markdown(option)

        engine.say(question)
        engine.say((options))
        engine.runAndWait()

        # Display the question on the Streamlit page
        #st.markdown(f"### Q{i+1}: {question}")
        #for option in options:
           # st.markdown(option)

        # Take input answer from microphone and convert it to text
        while True:
            with sr.Microphone() as source:
                st.write("Speak now...")

                # Record audio from the microphone
                audio = r.listen(source)

                # Convert speech to text
                try:
                    answer = r.recognize_google(audio)
                    st.write(f"You said: {answer}")

                    # Check if user wants to repeat the question
                    if "repeat the question" in answer.lower():
                        engine.say(question)
                        engine.say(options)

                        #engine.say(opti)
                        engine.runAndWait()
                        continue  # Ask the question again
                    #if user wants to repeat only particular part of a question
                    #if user wants to repeat only options
                    if "repeat options only" in answer.lower():
                        engine.say(options)
                        engine.runAndWait()
                        continue  # Ask the question again
                    #if user wants to repeat only option 1
                    if "repeat option one only" or  "repeat option 1" in answer.lower():
                        engine.say(options[0])
                        engine.runAndWait()
                        continue  # Ask the question again
                    #if user wants to repeat only option 2
                    if "repeat option two only" or  "repeat option 2" in answer.lower():
                        engine.say(options[1])
                        engine.runAndWait()
                        continue  # Ask the question again
                    #if user wants to repeat only option 3
                    if "repeat option three only" or  "repeat option 3" in answer.lower():
                        engine.say(options[2])
                        engine.runAndWait()
                        continue  # Ask the question again
                    #if user wants to repeat only option 4
                    if "repeat option four only" or "repeat option 4" in answer.lower():
                        engine.say(options[3])
                        engine.runAndWait()
                        continue  # Ask the question again
                    if "skip" in answer.lower():
                        #engine.say(question)
                        #engine.say(opti)
                        #engine.runAndWait()
                        continue  # Ask the question again
                    


                    # Check if user wants to end the exam
                    if "stop exam" in answer.lower():
                        wb.save('qa_data.xlsx')
                        st.warning("Exam ended.")
                        break  # End the exam
                    
                    # Store question and answer in Excel sheet
                    sheet.cell(row=row, column=1).value = lo
                    sheet.cell(row=row, column=2).value = answer
                    row += 1
                    lo=lo+1
                    break  # Move on to the next question

                except sr.UnknownValueError:
                    st.warning("Sorry, could not understand audio.")
                except sr.RequestError as e:
                    st.warning(f"Could not request results from Google Speech Recognition service: {e}")

        # If this is the last question or the user has ended the exam, save the Excel sheet and exit
        if i == len(questions) - 1 or "end exam" in answer.lower():
            wb.save('qa_data.xlsx')
            st.success("Exam completed.")
            break  # End





